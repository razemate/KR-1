import os
import chromadb
from chromadb.utils import embedding_functions
import PyPDF2
import openpyxl
import pandas as pd
from docx import Document
from typing import List, Dict, Any
import io

class RAGService:
    """RAG service for file processing and embeddings"""
    
    def __init__(self, storage_path: str = "../storage/chromadb"):
        self.storage_path = storage_path
        os.makedirs(storage_path, exist_ok=True)
        
        # Initialize ChromaDB
        self.client = chromadb.PersistentClient(path=storage_path)
        self.embedding_function = embedding_functions.DefaultEmbeddingFunction()
        
        # Get or create collection
        try:
            self.collection = self.client.get_collection(name="rag_documents")
        except:
            self.collection = self.client.create_collection(
                name="rag_documents",
                embedding_function=self.embedding_function
            )
    
    def process_file(self, file_content: bytes, filename: str, file_type: str) -> Dict[str, Any]:
        """Process uploaded file and extract text content"""
        try:
            text_content = ""
            
            if file_type == '.pdf':
                text_content = self._extract_pdf_text(file_content)
            elif file_type == '.docx':
                text_content = self._extract_docx_text(file_content)
            elif file_type == '.xlsx':
                text_content = self._extract_xlsx_text(file_content)
            elif file_type in ['.txt', '.md']:
                text_content = file_content.decode('utf-8')
            elif file_type == '.csv':
                text_content = self._extract_csv_text(file_content)
            else:
                return {"success": False, "error": f"Unsupported file type: {file_type}"}
            
            # Add to ChromaDB
            document_id = f"{filename}_{len(text_content)}"
            self.collection.add(
                documents=[text_content],
                metadatas=[{"filename": filename, "file_type": file_type}],
                ids=[document_id]
            )
            
            return {
                "success": True,
                "document_id": document_id,
                "text_length": len(text_content),
                "message": f"File {filename} processed and embedded successfully"
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    def _extract_pdf_text(self, file_content: bytes) -> str:
        """Extract text from PDF file"""
        pdf_file = io.BytesIO(file_content)
        reader = PyPDF2.PdfReader(pdf_file)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    
    def _extract_docx_text(self, file_content: bytes) -> str:
        """Extract text from DOCX file"""
        docx_file = io.BytesIO(file_content)
        doc = Document(docx_file)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    
    def _extract_xlsx_text(self, file_content: bytes) -> str:
        """Extract text from XLSX file"""
        xlsx_file = io.BytesIO(file_content)
        wb = openpyxl.load_workbook(xlsx_file)
        text = ""
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text += f"Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                text += row_text + "\n"
        return text
    
    def _extract_csv_text(self, file_content: bytes) -> str:
        """Extract text from CSV file"""
        csv_file = io.StringIO(file_content.decode('utf-8'))
        df = pd.read_csv(csv_file)
        return df.to_string()
    
    def query_documents(self, query: str, n_results: int = 5) -> str:
        """Query documents for relevant context"""
        try:
            results = self.collection.query(
                query_texts=[query],
                n_results=n_results
            )
            
            if results['documents'] and results['documents'][0]:
                # Combine relevant documents
                context = "\n\n".join(results['documents'][0])
                return context
            else:
                return ""
                
        except Exception as e:
            print(f"Error querying documents: {e}")
            return ""
    
    def list_documents(self) -> List[Dict[str, Any]]:
        """List all uploaded documents"""
        try:
            results = self.collection.get()
            documents = []
            
            for i, doc_id in enumerate(results['ids']):
                metadata = results['metadatas'][i] if results['metadatas'] else {}
                documents.append({
                    "id": doc_id,
                    "filename": metadata.get('filename', 'Unknown'),
                    "file_type": metadata.get('file_type', 'Unknown'),
                    "content_preview": results['documents'][i][:200] + "..." if len(results['documents'][i]) > 200 else results['documents'][i]
                })
            
            return documents
            
        except Exception as e:
            print(f"Error listing documents: {e}")
            return []
    
    def delete_document(self, document_id: str) -> bool:
        """Delete a document from the collection"""
        try:
            self.collection.delete(ids=[document_id])
            return True
        except Exception as e:
            print(f"Error deleting document: {e}")
            return False
    
    def clear_all_documents(self) -> bool:
        """Clear all documents from the collection"""
        try:
            # Delete the collection and recreate it
            self.client.delete_collection(name="rag_documents")
            self.collection = self.client.create_collection(
                name="rag_documents",
                embedding_function=self.embedding_function
            )
            return True
        except Exception as e:
            print(f"Error clearing documents: {e}")
            return False

# Global instance
rag_service = RAGService()